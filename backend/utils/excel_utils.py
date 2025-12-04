"""
Excel处理工具
Excel handling utilities
"""
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO


def create_application_template() -> BytesIO:
    """
    创建申报Excel模板
    Create application Excel template
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "申报信息"
    
    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 定义表头
    headers = [
        "申报单位名称",
        "项目名称",
        "项目负责人",
        "负责人职称",
        "团队成员",
        "项目摘要",
        "主要创新点",
        "技术指标",
        "应用价值",
        "经济效益",
        "社会效益",
        "联系人",
        "联系电话",
        "电子邮箱"
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 20
    
    # 添加示例数据
    example_data = [
        "XX科学技术研究院",
        "高性能铜合金材料关键技术研究及应用",
        "张三",
        "研究员",
        "李四,王五,赵六",
        "本项目针对高性能铜合金材料开发,解决了...",
        "1.创新性开发了...; 2.突破了...; 3.建立了...",
        "抗拉强度≥600MPa,导电率≥85%IACS",
        "可广泛应用于...,市场前景广阔",
        "近三年新增销售额5000万元,利润1000万元",
        "降低能耗20%,减少排放30%",
        "张三",
        "138xxxx1234",
        "zhangsan@example.com"
    ]
    
    for col, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_applications_to_excel(applications: List[Dict[str, Any]]) -> BytesIO:
    """
    导出申报列表到Excel
    Export applications to Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "申报列表"
    
    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 定义表头
    headers = [
        "序号",
        "项目名称",
        "申报单位",
        "项目负责人",
        "申报状态",
        "提交时间",
        "当前阶段",
        "评审平均分",
        "最终结果"
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入数据
    for idx, app in enumerate(applications, 1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=app.get('title', ''))
        ws.cell(row=idx+1, column=3, value=app.get('unit_name', ''))
        ws.cell(row=idx+1, column=4, value=app.get('leader_name', ''))
        ws.cell(row=idx+1, column=5, value=app.get('status', ''))
        ws.cell(row=idx+1, column=6, value=str(app.get('submission_time', '')))
        ws.cell(row=idx+1, column=7, value=app.get('current_stage', ''))
        
        # 评分
        score_summary = app.get('score_summary_json', {})
        avg_score = score_summary.get('average_score', '') if score_summary else ''
        ws.cell(row=idx+1, column=8, value=avg_score)
        
        ws.cell(row=idx+1, column=9, value=app.get('final_result', ''))
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_statistics_to_excel(statistics: Dict[str, Any]) -> BytesIO:
    """
    导出统计数据到Excel
    Export statistics to Excel
    """
    wb = Workbook()
    
    # 概览表
    ws1 = wb.active
    ws1.title = "统计概览"
    ws1['A1'] = "统计项目"
    ws1['B1'] = "数值"
    
    overview = statistics.get('overview', {})
    ws1['A2'] = "申报总数"
    ws1['B2'] = overview.get('total_applications', 0)
    ws1['A3'] = "组织总数"
    ws1['B3'] = overview.get('total_organizations', 0)
    ws1['A4'] = "专家总数"
    ws1['B4'] = overview.get('total_experts', 0)
    ws1['A5'] = "评审总数"
    ws1['B5'] = overview.get('total_reviews', 0)
    
    # 按状态分布
    if 'application_by_status' in statistics:
        ws2 = wb.create_sheet(title="申报状态分布")
        ws2['A1'] = "状态"
        ws2['B1'] = "数量"
        
        row = 2
        for status, count in statistics['application_by_status'].items():
            ws2.cell(row=row, column=1, value=status)
            ws2.cell(row=row, column=2, value=count)
            row += 1
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
